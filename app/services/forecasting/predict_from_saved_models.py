from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import holidays
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from prophet.serialize import model_from_json


def _to_bool(value) -> bool:
    if isinstance(value, bool):
        return value

    if pd.isna(value):
        return False

    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def _format_week_label(date_value, week_number: int) -> str:
    date_value = pd.Timestamp(date_value)
    return f"Sem {week_number}\n({date_value.strftime('%d %b')})"


def _write_styled_excel_report(
    forecast_df: pd.DataFrame,
    output_xlsx_path: Path,
    horizon_weeks: int,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Pronóstico"

    max_column = len(forecast_df.columns)
    title_fill = PatternFill("solid", fgColor="203B66")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    header_fill = PatternFill("solid", fgColor="203B66")
    total_fill = PatternFill("solid", fgColor="D9E2F3")
    row_fill = PatternFill("solid", fgColor="F2F2F2")
    white_font = Font(color="FFFFFF", bold=True)
    warning_font = Font(color="C00000", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=max_column,
    )
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = (
        f"Pronóstico de Demanda — Próximas {horizon_weeks} Semanas "
        "(V1 · solo demanda · revisión humana requerida)"
    )
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=max_column,
    )
    warning_cell = worksheet.cell(row=2, column=1)
    warning_cell.value = (
        "PRONÓSTICO, NO ORDEN DE COMPRA. Cantidades con valor exacto "
        "(decimales), SIN MOQ / empaque / inventario / tiempo de entrega. "
        "No comprar directamente de esta tabla — requiere revisión de un planeador."
    )
    warning_cell.fill = warning_fill
    warning_cell.font = warning_font
    warning_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[2].height = 42

    for column_index, column_name in enumerate(forecast_df.columns, start=1):
        cell = worksheet.cell(row=3, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    worksheet.row_dimensions[3].height = 42

    for row_index, row in enumerate(forecast_df.itertuples(index=False), start=4):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if column_index <= 5 else "right",
                vertical="center",
            )
            if row_index % 2 == 0:
                cell.fill = row_fill
            if column_index > 5:
                cell.number_format = '#,##0.00'
            if column_index == max_column:
                cell.fill = total_fill
                cell.font = Font(bold=True)

    for column_index, width in enumerate(
        [16, 48, 12, 22, 28] + [14] * (max_column - 6) + [16],
        start=1,
    ):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A4"
    worksheet.auto_filter.ref = (
        f"A3:{get_column_letter(max_column)}{len(forecast_df) + 3}"
    )

    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx_path)


def _build_weekly_holidays(
    country_code: str,
    start_year: int,
    end_year: int,
) -> pd.DataFrame:
    holiday_calendar = holidays.country_holidays(
        country_code,
        years=list(range(start_year, end_year + 1)),
    )

    rows = []

    for holiday_date, holiday_name in holiday_calendar.items():
        holiday_timestamp = pd.Timestamp(holiday_date)
        holiday_week_start = holiday_timestamp.to_period("W-SUN").start_time

        rows.append({
            "holiday": f"mx_{holiday_name}",
            "ds": holiday_week_start,
        })

    if not rows:
        return pd.DataFrame(columns=["holiday", "ds", "lower_window", "upper_window"])

    return (
        pd.DataFrame(rows)
        .drop_duplicates()
        .assign(lower_window=0, upper_window=0)
        .reset_index(drop=True)
    )


def _resolve_forecast_start_week(
    forecast_start_date: Optional[str],
) -> pd.Timestamp:
    resolved_date = forecast_start_date or date.today().isoformat()
    try:
        return pd.Timestamp(resolved_date).to_period("W-SUN").start_time
    except (TypeError, ValueError) as exc:
        raise ValueError(
            "forecast_start_date must use YYYY-MM-DD format."
        ) from exc


def run_saved_model_forecast(
    forecast_start_date: Optional[str] = None,
    horizon_weeks: Optional[int] = None,
    artifact_dir: Optional[str | Path] = None,
    output_dir: Optional[str | Path] = None,
    output_path: Optional[str | Path] = None,
) -> dict:
    """
    Load saved Prophet model artifacts and generate a top-N weekly sales forecast.

    If forecast_start_date is provided:
        forecast starts from that date's weekly start.

    If forecast_start_date is not provided:
        forecast starts from the week containing today's date.

    Output:
        A wide CSV with one row per product/UOM and one column per forecast week.
    """

    project_dir = Path(__file__).resolve().parents[3]

    artifact_dir = Path(artifact_dir) if artifact_dir else project_dir / "storage"
    output_path = Path(output_path) if output_path else None
    output_dir = (
        Path(output_dir)
        if output_dir
        else output_path.parent
        if output_path
        else project_dir / "storage" / "forecasts"
    )

    models_dir = artifact_dir / "models"
    registry_path = artifact_dir / "registry" / "model_registry.csv"
    metadata_path = artifact_dir / "metadata" / "training_metadata.json"

    output_dir.mkdir(parents=True, exist_ok=True)

    if not registry_path.exists():
        raise FileNotFoundError(f"Missing model registry: {registry_path}")

    if not metadata_path.exists():
        raise FileNotFoundError(f"Missing training metadata: {metadata_path}")

    with open(metadata_path, "r", encoding="utf-8") as f:
        metadata = json.load(f)

    model_registry = pd.read_csv(registry_path)

    if horizon_weeks is None:
        horizon_weeks = int(metadata.get("default_horizon_weeks", 4))
    if horizon_weeks < 1:
        raise ValueError("horizon_weeks must be at least 1.")

    training_last_week = metadata["training_last_week"]

    forecast_start_week = _resolve_forecast_start_week(
        forecast_start_date=forecast_start_date,
    )

    future_week_index = pd.date_range(
        start=forecast_start_week,
        periods=horizon_weeks,
        freq="7D",
    )

    forecast_end_week = future_week_index.max()

    holiday_country_code = metadata.get("holiday_country_code", "MX")
    training_first_week = metadata.get("training_first_week", training_last_week)

    weekly_holidays = _build_weekly_holidays(
        country_code=holiday_country_code,
        start_year=pd.Timestamp(training_first_week).year,
        end_year=forecast_end_week.year,
    )

    ma_blend_weight = float(metadata.get("ma_blend_weight", 0.30))

    forecast_rows = []
    errors = []

    for _, row in model_registry.iterrows():
        product_uom_id = row["product_uom_id"]
        model_file = row["model_file"]
        model_path = models_dir / model_file

        try:
            if not model_path.exists():
                raise FileNotFoundError(f"Missing model file: {model_path}")

            with open(model_path, "r", encoding="utf-8") as f:
                model = model_from_json(f.read())

            # Extend holidays to cover the requested forecast window.
            # This keeps prediction independent from raw training data.
            model.holidays = weekly_holidays

            use_lag = _to_bool(row.get("use_lag", False))
            ma_blended = _to_bool(row.get("ma_blended", False))

            future_df = pd.DataFrame({
                "ds": future_week_index,
            })

            if use_lag:
                lag_mean_value = float(row.get("lag_mean_value", 0.0))
                future_df["lag_mean"] = lag_mean_value

            forecast = model.predict(future_df)

            prophet_yhat = np.clip(
                forecast["yhat"].to_numpy(),
                a_min=0,
                a_max=None,
            )

            if ma_blended:
                ma_forecast_value = float(row.get("ma_forecast_value", 0.0))
                final_yhat = (
                    (1 - ma_blend_weight) * prophet_yhat
                    + ma_blend_weight * ma_forecast_value
                )
            else:
                final_yhat = prophet_yhat

            output_row = {
                "Código": row.get("product_id", ""),
                "Producto": row.get("product_name", ""),
                "UDM": row.get("uom", ""),
                "Categoría": row.get("category_l1", ""),
                "Subcategoría": row.get("category_l2", ""),
            }

            week_total = 0.0
            for week_number, (ds_value, yhat_value) in enumerate(
                zip(future_week_index, final_yhat),
                start=1,
            ):
                forecast_value = round(float(yhat_value), 2)
                output_row[_format_week_label(ds_value, week_number)] = forecast_value
                week_total += forecast_value

            output_row[f"Total {horizon_weeks} sem."] = round(week_total, 2)

            forecast_rows.append(output_row)

        except Exception as exc:
            errors.append({
                "product_uom_id": product_uom_id,
                "model_file": model_file,
                "error_message": str(exc),
            })

    wide_forecast_df = pd.DataFrame(forecast_rows)

    week_columns = [
        _format_week_label(date_value, week_number)
        for week_number, date_value in enumerate(future_week_index, start=1)
    ]

    expected_columns = [
        "Código",
        "Producto",
        "UDM",
        "Categoría",
        "Subcategoría",
    ] + week_columns
    expected_columns.append(f"Total {horizon_weeks} sem.")

    for col in expected_columns:
        if col not in wide_forecast_df.columns:
            wide_forecast_df[col] = ""

    wide_forecast_df = wide_forecast_df[expected_columns]

    start_label = forecast_start_week.strftime("%Y_%m_%d")

    output_csv_path = output_path or (
        output_dir
        / f"top_{len(wide_forecast_df)}_weekly_forecast_start_{start_label}_next_{horizon_weeks}_weeks_wide.csv"
    )

    error_csv_path = None

    if errors:
        error_df = pd.DataFrame(errors)
        error_csv_path = output_dir / f"forecast_errors_start_{start_label}.csv"
        error_df.to_csv(error_csv_path, index=False, encoding="utf-8-sig")
        raise RuntimeError(
            f"Forecast failed for {len(errors)} of {len(model_registry)} models. "
            f"See '{error_csv_path}'."
        )

    if wide_forecast_df.empty:
        raise RuntimeError("Forecast completed without generating any rows.")

    wide_forecast_df.to_csv(
        output_csv_path,
        index=False,
        encoding="utf-8-sig",
    )

    output_xlsx_path = output_csv_path.with_suffix(".xlsx")
    _write_styled_excel_report(
        forecast_df=wide_forecast_df,
        output_xlsx_path=output_xlsx_path,
        horizon_weeks=horizon_weeks,
    )

    return {
        "forecast_start_week": str(forecast_start_week.date()),
        "forecast_end_week": str(forecast_end_week.date()),
        "horizon_weeks": horizon_weeks,
        "forecast_rows": int(len(wide_forecast_df)),
        "output_csv_path": str(output_csv_path),
        "output_xlsx_path": str(output_xlsx_path),
        "error_count": int(len(errors)),
        "error_csv_path": str(error_csv_path) if error_csv_path else None,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--forecast-start-date",
        type=str,
        default=None,
        help="Optional forecast start date in YYYY-MM-DD format.",
    )

    parser.add_argument(
        "--horizon-weeks",
        type=int,
        default=None,
        help="Optional forecast horizon in weeks. Defaults to metadata value.",
    )

    parser.add_argument(
        "--artifact-dir",
        type=str,
        default=None,
        help="Optional artifact directory path.",
    )

    parser.add_argument(
        "--output-dir",
        type=str,
        default=None,
        help="Optional output directory path.",
    )

    parser.add_argument(
        "--output-path",
        type=str,
        default=None,
        help="Optional exact output CSV path.",
    )

    args = parser.parse_args()

    result = run_saved_model_forecast(
        forecast_start_date=args.forecast_start_date,
        horizon_weeks=args.horizon_weeks,
        artifact_dir=args.artifact_dir,
        output_dir=args.output_dir,
        output_path=args.output_path,
    )

    print(json.dumps(result, indent=2))
