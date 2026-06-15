from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from threading import Lock
from typing import Optional, Union
from zoneinfo import ZoneInfo
from zipfile import BadZipFile, ZipFile

import pandas as pd
from fastapi import status
from openpyxl import load_workbook

from app.core.config import get_settings
from app.services.file_storage_service import build_stored_filename
from app.services.forecast_service import (
    ForecastConfig,
    ForecastInputError,
    REQUIRED_COLUMNS,
    run_forecast_file,
)


YET_TO_START = "Yet to start"
LOADING = "Loading"
SUCCESSFUL = "Successful"
FAILED = "Failed"

WEEKLY_UPLOAD = "weekly data upload"
TRAIN_DATA_PREP = "train data prep"
FORECASTING = "forecasting"

PIPELINE_PHASES = (WEEKLY_UPLOAD, TRAIN_DATA_PREP, FORECASTING)
REQUIRED_XLSX_PARTS = {"[Content_Types].xml", "xl/workbook.xml"}
TRAINING_DATA_EXTENSIONS = {"csv", "xlsx"}


class PipelineError(ValueError):
    def __init__(
        self,
        message: str,
        code: str = "pipeline_error",
        status_code: int = status.HTTP_400_BAD_REQUEST,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.status_code = status_code


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


@dataclass
class PipelinePhase:
    phase: str
    status: str = YET_TO_START
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    error_message: Optional[str] = None


@dataclass
class PipelineSnapshot:
    phases: list[PipelinePhase]
    latest_upload_file_name: Optional[str]
    latest_upload_path: Optional[Path]
    latest_upload_url: Optional[str]
    train_data_path: Optional[Path]
    forecast_result_path: Optional[Path]
    latest_sales_date: Optional[str]


@dataclass
class ForecastPeriodInfo:
    training_start_date: str
    training_end_date: str
    forecasting_start_date: str
    forecasting_end_date: str
    forecast_horizon_weeks: int


class CGPipelineState:
    def __init__(self) -> None:
        self._lock = Lock()
        self._phases = {phase: PipelinePhase(phase=phase) for phase in PIPELINE_PHASES}
        self._latest_upload_file_name: Optional[str] = None
        self._latest_upload_path: Optional[Path] = None
        self._latest_upload_url: Optional[str] = None
        self._train_data_path: Optional[Path] = None
        self._forecast_result_path: Optional[Path] = None
        self._latest_sales_date: Optional[str] = None

    def snapshot(self) -> PipelineSnapshot:
        with self._lock:
            return PipelineSnapshot(
                phases=[
                    PipelinePhase(
                        phase=phase.phase,
                        status=phase.status,
                        started_at=phase.started_at,
                        completed_at=phase.completed_at,
                        error_message=phase.error_message,
                    )
                    for phase in self._phases.values()
                ],
                latest_upload_file_name=self._latest_upload_file_name,
                latest_upload_path=self._latest_upload_path,
                latest_upload_url=self._latest_upload_url,
                train_data_path=self._train_data_path,
                forecast_result_path=self._forecast_result_path,
                latest_sales_date=self._latest_sales_date,
            )

    def is_loading(self, phase_name: str) -> bool:
        with self._lock:
            return self._phases[phase_name].status == LOADING

    def mark_loading(self, phase_name: str) -> None:
        with self._lock:
            phase = self._phases[phase_name]
            phase.status = LOADING
            phase.started_at = utc_now_iso()
            phase.completed_at = None
            phase.error_message = None

    def mark_successful(self, phase_name: str) -> None:
        with self._lock:
            phase = self._phases[phase_name]
            phase.status = SUCCESSFUL
            phase.completed_at = utc_now_iso()
            phase.error_message = None

    def mark_failed(self, phase_name: str, error_message: str) -> None:
        with self._lock:
            phase = self._phases[phase_name]
            phase.status = FAILED
            phase.completed_at = utc_now_iso()
            phase.error_message = error_message

    def set_latest_upload(
        self,
        file_name: str,
        upload_path: Path,
        upload_url: str,
    ) -> None:
        with self._lock:
            self._latest_upload_file_name = file_name
            self._latest_upload_path = upload_path
            self._latest_upload_url = upload_url
            self._train_data_path = None
            self._forecast_result_path = None
            self._latest_sales_date = None
            self._phases[TRAIN_DATA_PREP] = PipelinePhase(phase=TRAIN_DATA_PREP)
            self._phases[FORECASTING] = PipelinePhase(phase=FORECASTING)

    def set_train_data_path(self, train_data_path: Path) -> None:
        with self._lock:
            self._train_data_path = train_data_path
            self._forecast_result_path = None
            self._latest_sales_date = read_latest_sales_date(train_data_path)
            self._phases[FORECASTING] = PipelinePhase(phase=FORECASTING)

    def mark_cg_data_prepared(self, train_data_path: Path) -> None:
        with self._lock:
            self._train_data_path = train_data_path
            self._forecast_result_path = None
            self._latest_sales_date = read_latest_sales_date(train_data_path)
            self._phases[FORECASTING] = PipelinePhase(phase=FORECASTING)

    def set_forecast_result_path(self, forecast_result_path: Path) -> None:
        with self._lock:
            self._forecast_result_path = forecast_result_path

    def try_mark_loading(self, phase_name: str) -> bool:
        with self._lock:
            phase = self._phases[phase_name]
            if phase.status == LOADING:
                return False
            phase.status = LOADING
            phase.started_at = utc_now_iso()
            phase.completed_at = None
            phase.error_message = None
            return True


pipeline_state = CGPipelineState()


def read_latest_sales_date(train_data_path: Path) -> Optional[str]:
    try:
        training_dates = pd.read_csv(train_data_path, usecols=["order_date"])[
            "order_date"
        ]
    except (FileNotFoundError, ValueError):
        return None

    parsed_dates = pd.to_datetime(training_dates, errors="coerce").dropna()
    if parsed_dates.empty:
        return None

    return parsed_dates.max().date().isoformat()


def validate_weekly_upload_content(filename: Optional[str], content: bytes) -> None:
    settings = get_settings()
    if not filename:
        raise PipelineError("Uploaded file must include a filename.", "missing_filename")

    extension = Path(filename).suffix.lower().lstrip(".")
    if extension not in settings.allowed_extensions:
        allowed = ", ".join(f".{item}" for item in settings.allowed_extensions)
        raise PipelineError(
            f"Only {allowed} files are supported.",
            "unsupported_file_type",
            status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
        )

    if not content:
        raise PipelineError("Uploaded file is empty.", "empty_file")

    if len(content) > settings.max_upload_size_bytes:
        max_mb = settings.max_upload_size_bytes // (1024 * 1024)
        raise PipelineError(
            f"Uploaded file must be {max_mb} MB or smaller.",
            "file_too_large",
            status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
        )

    try:
        with ZipFile(BytesIO(content)) as workbook:
            workbook_parts = set(workbook.namelist())
    except BadZipFile as exc:
        raise PipelineError(
            "The uploaded .xlsx file is not a valid Excel workbook.",
            "invalid_xlsx_file",
        ) from exc

    if not REQUIRED_XLSX_PARTS.issubset(workbook_parts):
        raise PipelineError(
            "The uploaded .xlsx file is missing required workbook data.",
            "invalid_xlsx_file",
        )

    validate_sales_sheet_headers(BytesIO(content), settings.source_train_file)


def validate_file_basics(
    filename: Optional[str],
    content: bytes,
    allowed_extensions: set[str],
) -> str:
    settings = get_settings()
    if not filename:
        raise PipelineError("Uploaded file must include a filename.", "missing_filename")

    extension = Path(filename).suffix.lower().lstrip(".")
    if extension not in allowed_extensions:
        allowed = ", ".join(f".{item}" for item in sorted(allowed_extensions))
        raise PipelineError(
            f"Only {allowed} files are supported.",
            "unsupported_file_type",
            status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
        )

    if not content:
        raise PipelineError("Uploaded file is empty.", "empty_file")

    if len(content) > settings.max_upload_size_bytes:
        max_mb = settings.max_upload_size_bytes // (1024 * 1024)
        raise PipelineError(
            f"Uploaded file must be {max_mb} MB or smaller.",
            "file_too_large",
            status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
        )

    return extension


def validate_xlsx_content(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as workbook:
            workbook_parts = set(workbook.namelist())
    except BadZipFile as exc:
        raise PipelineError(
            "The uploaded .xlsx file is not a valid Excel workbook.",
            "invalid_xlsx_file",
        ) from exc

    if not REQUIRED_XLSX_PARTS.issubset(workbook_parts):
        raise PipelineError(
            "The uploaded .xlsx file is missing required workbook data.",
            "invalid_xlsx_file",
        )


def get_sheet_headers(workbook_source: Union[Path, BytesIO], sheet_name: str) -> list[str]:
    workbook = load_workbook(workbook_source, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise PipelineError(
                f"Excel workbook must include sheet '{sheet_name}'.",
                "missing_sales_sheet",
            )

        worksheet = workbook[sheet_name]
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
    finally:
        workbook.close()

    return headers


def validate_sales_sheet_headers(upload_source: BytesIO, source_train_file: Path) -> None:
    sheet_name = ForecastConfig().excel_sheet_name
    upload_headers = get_sheet_headers(upload_source, sheet_name)
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in upload_headers]
    if missing_columns:
        missing = ", ".join(missing_columns)
        raise PipelineError(
            f"Uploaded workbook is missing required columns: {missing}",
            "missing_required_columns",
        )

    if not source_train_file.exists():
        return

    source_headers = get_sheet_headers(source_train_file, sheet_name)
    if upload_headers != source_headers:
        raise PipelineError(
            f"Uploaded workbook sheet '{sheet_name}' must match the existing train data columns.",
            "column_mismatch",
        )


def store_weekly_upload(filename: str, content: bytes) -> tuple[Path, str, str]:
    settings = get_settings()
    stored_filename = build_stored_filename(filename)
    settings.weekly_upload_dir.mkdir(parents=True, exist_ok=True)
    upload_path = settings.weekly_upload_dir / stored_filename
    upload_path.write_bytes(content)
    static_url = f"/static/uploads/{stored_filename}"
    return upload_path, stored_filename, static_url


def prepare_train_data() -> Path:
    settings = get_settings()
    snapshot = pipeline_state.snapshot()
    upload_path = snapshot.latest_upload_path
    if upload_path is None:
        raise PipelineError(
            "Upload weekly data before preparing train data.",
            "weekly_upload_required",
            status.HTTP_409_CONFLICT,
        )
    if not settings.source_train_file.exists():
        raise PipelineError(
            f"Source train file was not found: {settings.source_train_file}",
            "source_train_file_not_found",
            status.HTTP_404_NOT_FOUND,
        )

    sheet_name = ForecastConfig().excel_sheet_name
    source_sales = pd.read_excel(
        settings.source_train_file,
        sheet_name=sheet_name,
        usecols=REQUIRED_COLUMNS,
    )
    uploaded_sales = pd.read_excel(
        upload_path,
        sheet_name=sheet_name,
        usecols=REQUIRED_COLUMNS,
    )
    if uploaded_sales.empty:
        raise PipelineError(
            f"Uploaded workbook sheet '{sheet_name}' has no rows.",
            "empty_sales_sheet",
        )

    train_data = pd.concat([source_sales, uploaded_sales], ignore_index=True)
    settings.prepared_train_file.parent.mkdir(parents=True, exist_ok=True)
    train_data.to_csv(settings.prepared_train_file, index=False)
    pipeline_state.mark_cg_data_prepared(settings.prepared_train_file)
    return settings.prepared_train_file


def read_training_data(filename: Optional[str], content: bytes) -> pd.DataFrame:
    extension = validate_file_basics(filename, content, TRAINING_DATA_EXTENSIONS)

    if extension == "xlsx":
        validate_xlsx_content(content)
        workbook = pd.ExcelFile(BytesIO(content))
        if not workbook.sheet_names:
            raise PipelineError("Excel workbook has no sheets.", "missing_sales_sheet")

        sheet_name = (
            ForecastConfig().excel_sheet_name
            if ForecastConfig().excel_sheet_name in workbook.sheet_names
            else workbook.sheet_names[0]
        )
        columns = pd.read_excel(workbook, sheet_name=sheet_name, nrows=0).columns.tolist()
        missing_columns = [column for column in REQUIRED_COLUMNS if column not in columns]
        if missing_columns:
            missing = ", ".join(missing_columns)
            raise PipelineError(
                f"Training workbook is missing required columns: {missing}",
                "missing_required_columns",
            )
        training_data = pd.read_excel(
            workbook,
            sheet_name=sheet_name,
            usecols=REQUIRED_COLUMNS,
        )
    else:
        csv_source = BytesIO(content)
        columns = pd.read_csv(csv_source, nrows=0).columns.tolist()
        missing_columns = [column for column in REQUIRED_COLUMNS if column not in columns]
        if missing_columns:
            missing = ", ".join(missing_columns)
            raise PipelineError(
                f"Training CSV is missing required columns: {missing}",
                "missing_required_columns",
            )
        csv_source.seek(0)
        training_data = pd.read_csv(csv_source, usecols=REQUIRED_COLUMNS)

    if training_data.empty:
        raise PipelineError("Training data has no rows.", "empty_training_data")

    return training_data


def replace_training_data(filename: Optional[str], content: bytes) -> Path:
    settings = get_settings()
    training_data = read_training_data(filename, content)
    settings.prepared_train_file.parent.mkdir(parents=True, exist_ok=True)
    training_data.to_csv(settings.prepared_train_file, index=False)
    pipeline_state.set_train_data_path(settings.prepared_train_file)
    return settings.prepared_train_file


def get_forecast_period_info(
    train_data_path: Path,
    config: ForecastConfig = ForecastConfig(),
) -> ForecastPeriodInfo:
    try:
        training_dates = pd.read_csv(train_data_path, usecols=["order_date"])["order_date"]
    except ValueError as exc:
        raise PipelineError(
            "Training data is missing required column: order_date",
            "missing_required_columns",
        ) from exc

    parsed_dates = pd.to_datetime(training_dates, errors="coerce").dropna()
    if parsed_dates.empty:
        raise PipelineError(
            "Training data has no valid order_date values.",
            "invalid_training_dates",
        )

    training_start_date = parsed_dates.min().date().isoformat()
    training_end_date = parsed_dates.max().date().isoformat()
    latest_training_week = parsed_dates.dt.to_period("W-SUN").apply(
        lambda row: row.start_time
    ).max()

    forecasting_start_week = latest_training_week + pd.Timedelta(weeks=1)
    forecasting_end_week = latest_training_week + pd.Timedelta(
        weeks=config.forecast_horizon_weeks
    )
    forecasting_end_date = forecasting_end_week + pd.Timedelta(days=6)

    return ForecastPeriodInfo(
        training_start_date=training_start_date,
        training_end_date=training_end_date,
        forecasting_start_date=forecasting_start_week.date().isoformat(),
        forecasting_end_date=forecasting_end_date.date().isoformat(),
        forecast_horizon_weeks=config.forecast_horizon_weeks,
    )


def run_forecast_from_prepared_train() -> Path:
    settings = get_settings()
    snapshot = pipeline_state.snapshot()
    train_data_path = snapshot.train_data_path
    if train_data_path is None or not train_data_path.exists():
        raise PipelineError(
            "Prepare train data before running the forecast.",
            "train_data_required",
            status.HTTP_409_CONFLICT,
        )

    try:
        run_forecast_file(train_data_path, settings.forecast_result_file)
    except ForecastInputError:
        raise

    pipeline_state.set_forecast_result_path(settings.forecast_result_file)
    return settings.forecast_result_file


def run_forecast_pipeline_task() -> None:
    try:
        run_forecast_from_prepared_train()
    except Exception as exc:
        pipeline_state.mark_failed(FORECASTING, str(exc))
        return

    pipeline_state.mark_successful(FORECASTING)


def latest_sales_date_is_recent_enough(
    train_data_path: Path,
    max_age_weeks: int,
) -> bool:
    latest_sales_date = read_latest_sales_date(train_data_path)
    if latest_sales_date is None:
        return False

    settings = get_settings()
    latest_sales_day = date.fromisoformat(latest_sales_date)
    today = datetime.now(ZoneInfo(settings.scheduled_timezone)).date()
    age = today - latest_sales_day
    return age <= timedelta(weeks=max_age_weeks)


def run_scheduled_forecast_if_ready() -> bool:
    settings = get_settings()
    snapshot = pipeline_state.snapshot()
    if snapshot.train_data_path is None or not snapshot.train_data_path.exists():
        return False
    if not latest_sales_date_is_recent_enough(
        snapshot.train_data_path,
        settings.scheduled_forecast_max_sales_age_weeks,
    ):
        return False
    if not pipeline_state.try_mark_loading(FORECASTING):
        return False

    run_forecast_pipeline_task()
    return True
